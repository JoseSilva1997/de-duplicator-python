"""Reads .csv, .xlsx, and .xls files and returns a {sheet_name: SheetData} map.

`list_sheets` and `read` dispatch on file extension to format-specific helpers.
`_rows_to_sheet` finds the header row (skipping blank or title rows at the top),
then converts the remaining rows to ContactRecords, discarding any row that fails
ContactRecord.has_identifier().
"""
from __future__ import annotations

from pathlib import Path

from ..model import SheetData, ContactRecord, ContactField
from . import header_resolver
import csv
import openpyxl
import xlrd

def list_sheets(path: Path) -> list[str]:
    """Return the names of visible sheets. Hidden and very-hidden sheets are excluded.

    For CSV files, which have no sheet concept, returns a single synthetic name
    taken from the filename stem.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _list_sheets_csv(path)
    elif suffix == ".xlsx":
        return _list_sheets_xlsx(path)
    elif suffix == ".xls":
        return _list_sheets_xls(path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def read(path: Path, sheet_selection: list[str]) -> dict[str, SheetData]:
    """Read the given sheets and return a {sheet_name: SheetData} map in selection order.

    Column headers are resolved by header_resolver. Rows that fail
    ContactRecord.has_identifier() are dropped.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    elif suffix == ".xlsx":
        return _read_xlsx(path, sheet_selection)
    elif suffix == ".xls":
        return _read_xls(path, sheet_selection)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def _list_sheets_csv(path: Path) -> list[str]:
    # CSVs have no sheet concept, so the filename stem is used as a synthetic sheet name.
    return [path.stem]


def _read_csv(path: Path) -> dict[str, SheetData]:
    # sheet_selection is not threaded through here: a CSV always has exactly one sheet.
    # utf-8-sig strips the byte-order mark that Excel-exported CSVs often start with.
    # Without it, the first header cell picks up the BOM character and matches no alias.
    sheet_name = path.stem
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return {sheet_name: _rows_to_sheet(rows)}


def _list_sheets_xlsx(path: Path) -> list[str]:
    # read_only streams the workbook rather than loading it fully into memory.
    # data_only makes formula cells return their last cached value, not the formula text.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [name for name in wb.sheetnames if wb[name].sheet_state == "visible"]
    finally:
        wb.close()


def _read_xlsx(path: Path, sheet_selection: list[str]) -> dict[str, SheetData]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        result: dict[str, SheetData] = {}
        for name in sheet_selection:
            ws = wb[name]
            rows = [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
            result[name] = _rows_to_sheet(rows)
        return result
    finally:
        wb.close()


def _list_sheets_xls(path: Path) -> list[str]:
    wb = xlrd.open_workbook(str(path), on_demand=True)
    return [
        wb.sheet_names()[i]
        for i in range(wb.nsheets)
        if wb.sheet_by_index(i).visibility == 0  # xlrd: 0 = visible, 1 = hidden, 2 = very hidden
    ]


def _read_xls(path: Path, sheet_selection: list[str]) -> dict[str, SheetData]:
    wb = xlrd.open_workbook(str(path))
    result: dict[str, SheetData] = {}
    for name in sheet_selection:
        sheet = wb.sheet_by_name(name)
        rows = [
            [_cell_to_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        result[name] = _rows_to_sheet(rows)
    return result


def _cell_to_str(value) -> str | None:
    """Convert any cell value to a plain string, or None if the cell is empty.

    Handles the mix of types that openpyxl and xlrd produce (strings, floats,
    dates as floats, None). None is passed through intentionally so callers can
    distinguish a missing value from an empty string.
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _rows_to_sheet(rows: list[list[str]]) -> SheetData:
    # Scan from the top until a row resolves to a non-empty field map.
    # Blank rows and title rows with no ContactField aliases are skipped this way.
    field_map: dict[ContactField, int] = {}
    data_start = 0
    for i, row in enumerate(rows):
        field_map = header_resolver.resolve(row)
        if field_map:
            data_start = i + 1
            break
    records = []

    for row in rows[data_start:]:
        record = _row_to_record(row, field_map)
        if record is not None:
            records.append(record)

    return SheetData(records=records, available_fields=frozenset(field_map.keys()))


def _row_to_record(
        cells: list[str | None],
        field_map: dict[ContactField, int],
) -> ContactRecord | None:
    """Build a ContactRecord from one row, or return None if it has no usable identifier.

    The None return is the mechanism by which _rows_to_sheet silently drops
    structurally valid but content-empty rows (e.g. spacer rows in the sheet).
    """
    def get(field: ContactField) -> str | None:
        # Guard against ragged CSV rows where a trailing blank column makes the row
        # shorter than the header, which would otherwise cause an IndexError.
        idx = field_map.get(field)
        if idx is None or idx >= len(cells):
            return None
        return cells[idx]
    
    record = ContactRecord.build(
        first_name=get(ContactField.FIRST_NAME),
        last_name=get(ContactField.LAST_NAME),
        full_name=get(ContactField.FULL_NAME),
        email=get(ContactField.EMAIL),
        company=get(ContactField.COMPANY),
        job_title=get(ContactField.JOB_TITLE),
        country=get(ContactField.COUNTRY),
    )
    return record if record.has_identifier() else None